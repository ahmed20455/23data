from selenium import webdriver
from selenium.webdriver.common.by import By

# Set up the Selenium webdriver
driver = webdriver.Chrome()  # Update with the appropriate webdriver

# Open the website
driver.get('https://www.osmania.ac.in/res07/20230580.jsp')

# Open the result file in append mode
with open('result.txt', 'a') as file:
    # Iterate over roll numbers
    for roll_number in range(245322748001, 245322748128):
        # Find the input field and enter the roll number
        input_field = driver.find_element(By.NAME, 'htno')
        input_field.clear()
        input_field.send_keys(str(roll_number))

        # Submit the form
        driver.find_element(By.NAME, 'Submit').click()

        # Extract the text from the page
        page_text = driver.find_element(By.TAG_NAME, 'body').text

        # Extract lines 4 and 19
        lines = page_text.split('\n')
        filtered_lines = [lines[3]]

        # Extract the last character from lines 9 to 16
        last_characters = [line[-1] for line in lines[8:16]]

        # Concatenate lines 4, 19, and last characters from lines 9 to 16
        filtered_text = '\n'.join(filtered_lines + last_characters)

        # Append the filtered text to the result file
        file.write(filtered_text + '\n')

# Close the browser
driver.quit()

import pandas as pd

# Specify the number of lines to skip after each range of rows
lines_to_skip = 1

# Read the text file
with open('result.txt', 'r') as file:
    lines = file.readlines()

# Initialize variables
data = []
start_line = 1
end_line = 9

# Process the text file data
while start_line < len(lines):
    # Extract the desired lines
    extracted_lines = lines[start_line:end_line]

    # Append the extracted lines to the data list
    data.append(extracted_lines)

    # Update the start and end lines for the next iteration
    start_line = end_line + lines_to_skip
    end_line = start_line + 8

# Convert the data to a pandas DataFrame
df = pd.DataFrame(data)

# Save the DataFrame to an Excel file
df.to_excel('output.xlsx', index=False, header=False)


import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('output.xlsx')

# Select the desired sheet
sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name
sheet.insert_cols(1)
sheet.insert_rows(1)
roll_number = "Roll Numbers"  # Replace with the actual roll number
sheet['A1'] = roll_number

# Generate and write roll numbers in cells A1 to A60
start_roll = 245322748001
for i in range(127):
    roll_number = str(start_roll + i)
    cell = sheet.cell(row=i+2, column=1)
    cell.value = roll_number

# Save the modified Excel file
workbook.save('output.xlsx')

# Load the Excel file
workbook = openpyxl.load_workbook('output.xlsx')

# Select the desired sheet
sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name

# Insert a column at the leftmost side

# Write roll number in cell A1
sheet['B1'] = 'M1'
sheet['C1'] = 'PHY'
sheet['D1'] = 'BEE'
sheet['E1'] = 'PHY L'
sheet['F1'] = 'CMS'
sheet['G1'] = 'IC'
sheet['H1'] = 'EGDP'
sheet['I1'] = 'ENG L'
sheet['J1'] = 'SUB FAILED'

# Save the modified Excel file
workbook.save('output.xlsx')

import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('output.xlsx')

# Select the desired sheet
sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name

# Iterate over the rows 2 to 61
for row in range(2, 128):
    count = 0

    # Iterate over the columns B to I
    for column in range(2, 10):
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value and isinstance(cell_value, str) and 'F' in cell_value:
            count += 1

    # Store the count in the respective cell in column J
    sheet.cell(row=row, column=10).value = count

# Save the modified Excel file
workbook.save('output.xlsx')


# Load the Excel file
workbook = openpyxl.load_workbook('output.xlsx')

# Select the desired sheet
sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name

# Initialize the sum variable
column_j_sum = 0

# Iterate over the cells in column J from row 2 to row 61
for row in range(2, 128):
    cell_value = sheet.cell(row=row, column=10).value
    if cell_value and isinstance(cell_value, int):
        column_j_sum += cell_value

# Store the sum in cell J62
sheet['J128'] = column_j_sum

# Save the modified Excel file
workbook.save('output.xlsx')


# Load the Excel file
workbook = openpyxl.load_workbook('output.xlsx')

# Select the desired sheet
sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name

# Iterate over the columns B to I
for column in range(2, 10):
    # Initialize the count variable for each column
    count = 0

    # Iterate over the cells in the column from row 2 to row 61
    for row in range(2, 128):
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value and isinstance(cell_value, str) and 'F' in cell_value:
            count += 1

    # Store the count in the 62nd cell of the respective column
    sheet.cell(row=128, column=column).value = count

# Save the modified Excel file
workbook.save('output.xlsx')
